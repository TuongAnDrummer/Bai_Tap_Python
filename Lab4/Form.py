# Import openpyxl and Tkinter modules
from openpyxl import load_workbook
import tkinter as tk
from tkinter import W, Checkbutton, ttk
import tkinter.messagebox as mbox
import re
# Globally declare wb and sheet variable

# Opening the existing excel file
wb = load_workbook('excel.xlsx')

# Create the sheet object
sheet = wb.active


class Form:
    def __init__(self, root):
        self.root = root
        self.root.configure(background='light green')
        self.root.title("registration form")
        self.root.geometry("600x300")
        self.current_var = tk.StringVar()

        self.value1 = tk.StringVar()
        self.value2 = tk.StringVar()
        self.value3 = tk.StringVar()
        self.value4 = tk.StringVar()

        self.excel()

        # Create a Form label
        self.heading = tk.Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green",font=("Arial", 20),fg="red")

        # Create labels for each input field
        self.iD = tk.Label(root, text="Mã số sinh viên", bg="light green")
        self.hoTen = tk.Label(root, text="Họ tên", bg="light green")
        self.ngaySinh = tk.Label(root, text="Ngày sinh", bg="light green")
        self.email = tk.Label(root, text="Email", bg="light green")
        self.SDT = tk.Label(root, text="Số điện thoại", bg="light green")
        self.hocKy = tk.Label(root, text="Học kỳ", bg="light green")
        self.namHoc = tk.Label(root, text="Năm học", bg="light green")
        self.chonMonHoc = tk.Label(root, text="Chọn môn học", bg="light green")
        self.l1 = Checkbutton(root,
                text='Lập trình Python',
                variable= self.value1,
                onvalue='Lập trình Python',
                offvalue='None',
                bg="light green")
        self.l2 = Checkbutton(root,
                text='Lập trình Java',
                variable= self.value2,
                onvalue='Lập trình Java',
                offvalue='None',
                bg="light green")

        self.r1 = Checkbutton(root,
                text='Công nghệ phần mềm',
                variable= self.value3,
                onvalue='Công nghệ phần mềm',
                offvalue='None',
                bg="light green")
        self.r2 = Checkbutton(root,
                text='Phát triển ứng dụng web',
                variable= self.value4,
                onvalue='Phát triển ứng dụng web',
                offvalue='None',
                bg="light green")

        # Grid method is used for placing the widgets at respective positions
        self.heading.grid(row=0, column=1)
        self.iD.grid(row=1, column=0, sticky = W)
        self.hoTen.grid(row=2, column=0, sticky = W)
        self.ngaySinh.grid(row=3, column=0, sticky = W)
        self.email.grid(row=4, column=0, sticky = W)
        self.SDT.grid(row=5, column=0, sticky = W)
        self.hocKy.grid(row=6, column=0, sticky = W)
        self.namHoc.grid(row=7, column=0, sticky = W)
        self.chonMonHoc.grid(row=8,column=0, sticky = W)

        self.l1.grid(row = 8, column = 1, sticky = W,padx=50)
        self.l2.grid(row = 9, column = 1, sticky = W,padx=50)
        self.r1.grid(row = 8, column = 2, sticky = W)
        self.r1.place(x=300,y=185)
        self.r2.grid(row = 9, column = 2, sticky = W)
        self.r2.place(x=300,y=210)


        # Create text entry boxes for each input field
        self.iD_field = tk.Entry(root,validate="key",validatecommand=(root.register(validate_entry),"%P",7))
        self.hoTen_field = tk.Entry(root)
        self.ngaySinh_field = tk.Entry(root)
        self.email_field = tk.Entry(root)
        self.SDT_field = tk.Entry(root,validate="key",validatecommand=(root.register(validate_entry), "%P",10))
        self.hocKy_field = tk.Entry(root,validate="key",validatecommand=(root.register(validate_entry_HocKy), "%P",1))
        self.namHoc_field = ttk.Combobox(root,textvariable=self.current_var)
        self.namHoc_field['values'] = ('2022-2023', '2023-2024', '2024-2025')        

        # Bind the Enter key to focus functions
        self.hoTen_field.bind("<Return>", self.focus1)
        self.ngaySinh_field.bind("<Return>", self.focus2)
        self.email_field.bind("<Return>", self.focus3)
        self.SDT_field.bind("<Return>", self.focus4)
        self.hocKy_field.bind("<Return>", self.focus5)
        self.namHoc_field.bind("<Return>", self.focus6)

        # Grid method is used for placing the text entry boxes
        self.iD_field.grid(row=1, column=1, ipadx="100")
        self.hoTen_field.grid(row=2, column=1, ipadx="100")
        self.ngaySinh_field.grid(row=3, column=1, ipadx="100")
        self.email_field.grid(row=4, column=1, ipadx="100")
        self.SDT_field.grid(row=5, column=1, ipadx="100")
        self.hocKy_field.grid(row=6, column=1, ipadx="100")
        self.namHoc_field.grid(row=7, column=1, ipadx="91")
        

        # Create a Submit Button and place it into the root window
        self.submit = tk.Button(root, text="Đăng Ký", fg="Black", bg="light green", command=self.insert)
        self.submit.grid(row=10, column=1)

    # Function to resize the width of columns in the excel spreadsheet

    def excel(self):
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 50


        # Write given data to an excel spreadsheet at a particular location
        sheet.cell(row=1, column=1).value = "ID"
        sheet.cell(row=1, column=2).value = "Họ Tên"
        sheet.cell(row=1, column=3).value = "Ngày Sinh"
        sheet.cell(row=1, column=4).value = "Email"
        sheet.cell(row=1, column=5).value = "Số điện thoại"
        sheet.cell(row=1, column=6).value = "Học kỳ"
        sheet.cell(row=1, column=7).value = "Năm học"
        sheet.cell(row=1, column=8).value = "Chọn môn học"


    # Function to set focus (cursor) on the course_field box
    def focus1(self, event):
        self.hoTen_field.focus_set()

    # Function to set focus on the sem_field box
    def focus2(self, event):
        self.ngaySinh_field.focus_set()

    # Function to set focus on the email_field box
    def focus3(self, event):
        self.email_field.focus_set()

    # Function to set focus on the contact_no_field box
    def focus4(self, event):
        self.SDT_field.focus_set()

    # Function to set focus on the email_id_field box
    def focus5(self, event):
        self.hocKy_field.focus_set()

    # Function to set focus on the address_field box
    def focus6(self, event):
        self.namHoc_field.focus_set()

    # Function for clearing the contents of text entry boxes
    def clear(self):
        self.iD_field.delete(0, tk.END)
        self.hoTen_field.delete(0, tk.END)
        self.ngaySinh_field.delete(0, tk.END)
        self.SDT_field.delete(0, tk.END)
        self.email_field.delete(0, tk.END)
        self.hocKy_field.delete(0, tk.END)
        self.namHoc_field.delete(0, tk.END)
        self.l1.deselect()
        self.l2.deselect()
        self.r1.deselect()
        self.r2.deselect()

    # Function to take data from GUI window and write to an excel file
    def insert(self):
        # If user did not fill any entry, print "empty input"
        if (self.iD_field.get() == "" and
            self.hoTen_field.get() == "" and
            self.ngaySinh_field.get() == "" and
            self.SDT_field.get() == "" and
            self.email_field.get() == "" and
            self.hocKy_field.get() == "" and
            self.namHoc_field.get() == ""):
            mbox.showwarning("Warning","Các ô đang để trống")
        else:
            # Assign the max row and max column value to the variables
            current_row = sheet.max_row
            current_column = sheet.max_column
			
            if len(str(self.iD_field.get())) < 7 :
                mbox.showwarning("Warning","Mã số sinh viên gồm 7 số")
            elif(ValidEmail(self.email_field.get()) == 0):
                mbox.showwarning("Warning","Email khong hop le")	
            elif(len(str(self.SDT_field.get())) < 10 ):
                mbox.showwarning("Warning","Số điện thoại gồm 10-11 số")
            elif(ValidDate(self.ngaySinh_field.get()) == 0):
                mbox.showwarning("Warning","Ngay sinh khong hop le")
            else:
                
                sheet.cell(row=current_row + 1, column=1).value = self.iD_field.get()
                sheet.cell(row=current_row + 1, column=2).value = self.hoTen_field.get()
                sheet.cell(row=current_row + 1, column=3).value = self.ngaySinh_field.get()
                sheet.cell(row=current_row + 1, column=4).value = self.SDT_field.get()
                sheet.cell(row=current_row + 1, column=5).value = self.email_field.get()
                sheet.cell(row=current_row + 1, column=6).value = self.hocKy_field.get()
                sheet.cell(row=current_row + 1, column=7).value = self.namHoc_field.get()
                sheet.cell(row=current_row + 1, column=8).value = self.value1.get()
            
                
                # sheet.cell(row=current_row + 1, column=8).value = self.value2.get()
                # sheet.cell(row=current_row + 1, column=8).value = self.value3.get()
                # sheet.cell(row=current_row + 1, column=8).value = self.value4.get()

                # Save the file
                wb.save('excel.xlsx')

                # Set focus on the name_field box
                self.iD_field.focus_set()

                # Call the clear() function
                self.clear()
############################################################           
# def validate_entry_MS(text):
#     return len(text) == 0 or text.isdecimal() and len(text) <= 7

# def validate_entry_SDT(text):
#     return len(text) == 0 or text.isdecimal() and len(text) < 10

# def validate_entry_HocKy(text):
#     return len(text) == 0 or text.isdecimal() and len(text) <= 1

def validate_entry(text,n:int):
    return len(text) == 0 or text.isdecimal() and len(text) <= int(n)

def validate_entry_HocKy(text,n:int):
    return len(text) == 0 or text.isdecimal() and len(text) <= int(n) and int(text) <= 3

def ValidEmail(string):
    return bool(re.match(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',string))

def ValidDate(string):
    return bool(re.match(r'^\d{2}/\d{2}/\d{4}$',string))


if __name__ == "__main__":
    # Create a GUI window
    root = tk.Tk()
    form = Form(root)
    root.mainloop()
